import openpyxl

class ExcelUtils:
    def __init__(self, file, sheetname, email_column=1, header_row=1):
        """
        :param file: Path to the Excel file
        :param sheetname: Name of the sheet (e.g., 'Sheet1', 'Sheet2')
        :param email_column: Column containing emails, default is column 1 (A), can specify other columns
        :param header_row: Row containing column headers, default is row 1
        """
        self.file = file
        self.sheetname = sheetname
        self.email_column = email_column
        self.header_row = header_row  # Column header row
        self.workload = openpyxl.load_workbook(file)
        self.sheet = self.workload[sheetname]

    def get_row_count(self):
        """Returns the number of rows in the sheet"""
        return self.sheet.max_row

    def get_column_count(self):
        """Returns the number of columns in the sheet"""
        return self.sheet.max_column

    def read_data(self, row, col):
        """Reads data at position (row, col)"""
        return self.sheet.cell(row, col).value

    def check_email_exist(self, email):
        """Checks if the email exists in the email column, ignoring the column header"""
        # Start from the row containing data, after the header row (header_row + 1)
        start_row = self.header_row + 1
        for r in range(start_row, self.get_row_count() + 1):
            if self.read_data(r, self.email_column) == email:
                return True
        return False
